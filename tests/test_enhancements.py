from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from factor_risk_analysis.config import DEFAULT_FACTORS
from factor_risk_analysis.enhancements import enhance_workbook
from factor_risk_analysis.workbook import WorkbookSettings, create_workbook


def make_sample_frame(months: int = 36) -> pd.DataFrame:
    idx = pd.date_range("2021-01-31", periods=months, freq="ME")
    data = pd.DataFrame(index=idx)
    data["Fund"] = [0.02, -0.03, -0.02, 0.01, 0.04, -0.01] * 6
    data["Benchmark"] = [0.015, -0.02, -0.01, 0.015, 0.03, -0.005] * 6
    data["Equity"] = [0.01 + ((i % 6) - 3) * 0.004 for i in range(months)]
    data["Currency"] = [0.001 + ((i % 7) - 3) * 0.002 for i in range(months)]
    data["Credit"] = [0.003 + ((i % 5) - 2) * 0.002 for i in range(months)]
    data["Volatility"] = [-0.002 + ((i % 8) - 4) * 0.01 for i in range(months)]
    data["VIX Level"] = [35 if i % 10 == 0 else 20 for i in range(months)]
    return data


def test_enhanced_workbook_adds_guidance_and_drawdown_duration():
    data = make_sample_frame()
    settings = WorkbookSettings(
        fund_name="Test Fund",
        input_mode="Unit test",
        fund_source="Synthetic",
        requested_start="2021-01-01",
        requested_end="2023-12-31",
        risk_free_rate_annual=0.03,
        factors=DEFAULT_FACTORS,
        benchmark_name="SPY",
    )
    payload = enhance_workbook(create_workbook(data, settings))
    wb = load_workbook(BytesIO(payload), data_only=False)

    assert wb["Regression"]["A6"].value == "Coefficients"
    assert wb["Regression"]["B4"].value == "Volatility"
    assert wb["Conditional Regression"]["B4"].value == "Crisis x Volatility"
    assert wb["Fund Returns"]["N5"].value == "=IF(F5<0,1,0)"
    assert wb["Risk Statistics"]["B24"].value.startswith("=MAX('Fund Returns'!$N$")
    assert wb["Summary"]["B32"].value == "='Risk Statistics'!B24"
    assert any(cell.value == "How to Interpret Diagnostics" for row in wb["Diagnostics"].iter_rows() for cell in row)
    assert any(cell.value == "How to Interpret Rolling Analysis" for row in wb["Rolling Analysis"].iter_rows() for cell in row)
    assert any(cell.value == "How to Interpret the F-statistic" for row in wb["Summary"].iter_rows() for cell in row)
