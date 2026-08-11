from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from factor_risk_analysis.config import DEFAULT_FACTORS
from factor_risk_analysis.workbook import WorkbookSettings, create_workbook


def make_sample_frame(months: int = 36) -> pd.DataFrame:
    idx = pd.date_range("2021-01-31", periods=months, freq="ME")
    data = pd.DataFrame(index=idx)
    data["Fund"] = [0.005 + ((i % 5) - 2) * 0.002 for i in range(months)]
    data["Benchmark"] = [0.004 + ((i % 4) - 1.5) * 0.003 for i in range(months)]
    data["Equity"] = [0.006 + ((i % 6) - 3) * 0.004 for i in range(months)]
    data["Currency"] = [0.001 + ((i % 7) - 3) * 0.002 for i in range(months)]
    data["Credit"] = [0.003 + ((i % 5) - 2) * 0.002 for i in range(months)]
    data["Volatility"] = [-0.002 + ((i % 8) - 4) * 0.01 for i in range(months)]
    data["VIX Level"] = [20 + (15 if i % 10 == 0 else 0) for i in range(months)]
    return data


def test_workbook_contains_expected_sheets_and_linest_formulas():
    data = make_sample_frame()
    settings = WorkbookSettings(
        fund_name="Test Fund",
        input_mode="Unit test",
        fund_source="Synthetic",
        requested_start="2021-01-01",
        requested_end="2023-12-31",
        risk_free_rate_annual=0.03,
        factors=DEFAULT_FACTORS,
        benchmark_name="SPY",
    )
    payload = create_workbook(data, settings)
    path = "/tmp/test_factor_risk_analysis.xlsx"
    with open(path, "wb") as f:
        f.write(payload)

    wb = load_workbook(path, data_only=False)
    expected = {
        "Summary",
        "Fund Returns",
        "Factor Data",
        "Regression",
        "Conditional Regression",
        "Risk Statistics",
        "Diagnostics",
        "Rolling Analysis",
        "Definitions",
    }
    assert expected.issubset(set(wb.sheetnames))
    assert wb["Regression"].array_formulae["B6"] == "B6:F10"
    assert wb["Conditional Regression"].array_formulae["B6"] == "B6:J10"
    assert wb["Risk Statistics"]["B21"].value.startswith("=IF(COUNT")
