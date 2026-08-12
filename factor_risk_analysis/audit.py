"""Post-processing helpers for workbook auditability."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from .config import SHEET_DIAGNOSTICS
from .credit_guidance import clarify_credit_factor

HEADER_FILL = "1F4E78"
THIN_GRAY = Side(style="thin", color="D9E2F3")
DATE_FMT = "mmm-yy"


def add_dropped_month_audit(workbook_bytes: bytes, dropped_details: pd.DataFrame) -> bytes:
    """Apply credit-factor guidance, then append an explicit dropped-month table."""

    workbook_bytes = clarify_credit_factor(workbook_bytes)
    wb = load_workbook(BytesIO(workbook_bytes))
    ws = wb[SHEET_DIAGNOSTICS]

    # Keep this section below the existing warning/guidance blocks while leaving
    # room for variable-length explanatory content.
    start_row = max(45, ws.max_row + 3)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    title = ws.cell(start_row, 1, "Dropped Month Detail")
    title.font = Font(bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=HEADER_FILL)

    header_row = start_row + 1
    headers = ["Date", "Missing series"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

    if dropped_details is None or dropped_details.empty:
        ws.cell(header_row + 1, 1, "No months were dropped for missing required series.")
    else:
        for offset, (_, row) in enumerate(dropped_details.iterrows(), start=1):
            excel_row = header_row + offset
            date_value = pd.Timestamp(row["Date"]).to_pydatetime()
            ws.cell(excel_row, 1, date_value)
            ws.cell(excel_row, 1).number_format = DATE_FMT
            ws.cell(excel_row, 2, str(row["Missing series"]))
            for col in (1, 2):
                ws.cell(excel_row, col).border = Border(bottom=THIN_GRAY)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 28)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 60)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
