"""Clarify the economic interpretation of the HYG credit proxy in generated workbooks."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .config import (
    SHEET_CONDITIONAL,
    SHEET_DEFINITIONS,
    SHEET_DIAGNOSTICS,
    SHEET_FACTORS,
    SHEET_REGRESSION,
    SHEET_ROLLING,
    SHEET_SUMMARY,
)

HEADER_FILL = "1F4E78"
SUBHEADER_FILL = "D9EAF7"
STATIC_GRAY = "666666"
THIN_GRAY = Side(style="thin", color="D9E2F3")

DISPLAY_NAME = "Credit (HYG Total Return Proxy)"
CRISIS_DISPLAY_NAME = "Crisis x Credit (HYG)"
CREDIT_NOTE = (
    "Credit is proxied by HYG monthly total return, not by the percentage change in a credit yield or spread. "
    "Holding the other model factors constant, a positive Credit beta means the fund tends to move in the same "
    "direction as HYG: HYG gains contribute positively to predicted fund return and HYG losses contribute negatively. "
    "A negative Credit beta indicates hedge-like or inverse exposure to HYG. Because HYG returns also reflect Treasury "
    "rates/duration, coupon carry, liquidity and other effects, this coefficient should be interpreted as high-yield "
    "credit-market exposure rather than a pure credit-spread beta."
)


def _section_title(ws, row: int, first_col: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    cell = ws.cell(row, first_col, text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def _write_credit_guide(ws, start_row: int, first_col: int = 1, last_col: int = 8) -> None:
    _section_title(ws, start_row, first_col, last_col, "How to Interpret the Credit Factor")
    rows = [
        ("Proxy", "HYG monthly total return is used as the credit-risk proxy. It is not a yield-change or spread-change series."),
        ("Positive beta", "The fund tends to participate with HYG. If HYG performs poorly, the model predicts a negative credit contribution to the fund, all else equal. A large positive beta therefore indicates substantial pro-credit / high-yield credit-market exposure."),
        ("Negative beta", "The fund tends to move opposite HYG. HYG weakness would contribute positively to predicted fund return, all else equal; this is hedge-like or short-credit behavior."),
        ("Magnitude", "A beta of 1.5 means a 1.0% HYG monthly return is associated with about a 1.5% modeled contribution from the credit factor, holding Equity, Currency and Volatility constant."),
        ("Important caveat", "HYG mixes credit-spread risk with Treasury-rate/duration exposure, coupon carry, liquidity and other effects. Treat the coefficient as HYG/high-yield credit-market exposure, not as a pure spread beta."),
        ("Conditional model", "Normal Credit beta is the non-crisis exposure. Effective crisis Credit beta equals Normal Credit beta plus Crisis Incremental Credit beta."),
    ]
    for offset, (label, text) in enumerate(rows, start=1):
        row = start_row + offset
        ws.cell(row, first_col, label)
        ws.cell(row, first_col).font = Font(bold=True, color=STATIC_GRAY)
        ws.cell(row, first_col).fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
        ws.merge_cells(start_row=row, start_column=first_col + 1, end_row=row, end_column=last_col)
        cell = ws.cell(row, first_col + 1, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(first_col, last_col + 1):
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)
        ws.row_dimensions[row].height = 44


def clarify_credit_factor(workbook_bytes: bytes) -> bytes:
    """Relabel and document the HYG credit factor without changing model calculations."""

    wb = load_workbook(BytesIO(workbook_bytes))

    # Factor data sheet.
    ws = wb[SHEET_FACTORS]
    ws["D4"] = DISPLAY_NAME
    ws["D4"].comment = Comment(CREDIT_NOTE, "Factor Risk Analysis")

    # Standard regression: LINEST label and coefficient term.
    ws = wb[SHEET_REGRESSION]
    ws["C4"] = DISPLAY_NAME
    ws["A17"] = DISPLAY_NAME
    ws["A17"].comment = Comment(CREDIT_NOTE, "Factor Risk Analysis")
    _write_credit_guide(ws, max(ws.max_row + 2, 45), 1, 9)

    # Conditional regression: both the normal and interaction columns use HYG returns.
    ws = wb[SHEET_CONDITIONAL]
    ws["C4"] = CRISIS_DISPLAY_NAME
    ws["G4"] = DISPLAY_NAME
    ws["A17"] = DISPLAY_NAME
    ws["A17"].comment = Comment(CREDIT_NOTE, "Factor Risk Analysis")
    _write_credit_guide(ws, max(ws.max_row + 2, 45), 1, 10)

    # Diagnostics labels.
    ws = wb[SHEET_DIAGNOSTICS]
    ws["E4"] = DISPLAY_NAME
    ws["A8"] = DISPLAY_NAME
    ws["D13"] = DISPLAY_NAME
    ws["H13"] = CRISIS_DISPLAY_NAME
    ws["A16"] = DISPLAY_NAME
    ws["A20"] = CRISIS_DISPLAY_NAME

    # Rolling beta label.
    ws = wb[SHEET_ROLLING]
    ws["D4"] = DISPLAY_NAME
    ws["D4"].comment = Comment(CREDIT_NOTE, "Factor Risk Analysis")

    # Definitions sheet: make the proxy explicit where the factor is listed.
    ws = wb[SHEET_DEFINITIONS]
    for row in range(19, 30):
        if ws.cell(row, 1).value == "Credit":
            ws.cell(row, 1, DISPLAY_NAME)
            ws.cell(row, 4, "High-yield credit-market risk proxy; HYG monthly total return. Positive beta means the fund tends to participate with HYG, all else equal.")
            ws.cell(row, 1).comment = Comment(CREDIT_NOTE, "Factor Risk Analysis")
            break
    _write_credit_guide(ws, max(ws.max_row + 2, 48), 1, 8)

    # Summary: include a compact sign-convention reminder.
    ws = wb[SHEET_SUMMARY]
    start_row = max(ws.max_row + 2, 58)
    _section_title(ws, start_row, 1, 5, "Credit Factor Sign Convention")
    ws.cell(start_row + 1, 1, "Positive Credit beta")
    ws.cell(start_row + 1, 1).font = Font(bold=True, color=STATIC_GRAY)
    ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=5)
    ws.cell(start_row + 1, 2, "Pro-HYG exposure: HYG weakness is expected to hurt the fund, all else equal. Larger positive beta implies greater high-yield credit-market sensitivity.")
    ws.cell(start_row + 2, 1, "Negative Credit beta")
    ws.cell(start_row + 2, 1).font = Font(bold=True, color=STATIC_GRAY)
    ws.merge_cells(start_row=start_row + 2, start_column=2, end_row=start_row + 2, end_column=5)
    ws.cell(start_row + 2, 2, "Inverse / hedge-like HYG exposure: HYG weakness is expected to help the fund, all else equal.")
    ws.cell(start_row + 3, 1, "Caveat")
    ws.cell(start_row + 3, 1).font = Font(bold=True, color=STATIC_GRAY)
    ws.merge_cells(start_row=start_row + 3, start_column=2, end_row=start_row + 3, end_column=5)
    ws.cell(start_row + 3, 2, "This is an HYG total-return beta, not a pure credit-spread beta; HYG also contains rate/duration, carry and liquidity effects.")
    for row in range(start_row + 1, start_row + 4):
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 38
        for col in range(1, 6):
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
