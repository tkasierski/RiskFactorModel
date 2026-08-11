"""Workbook presentation and interpretation enhancements."""

from __future__ import annotations

from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .config import (
    SHEET_CONDITIONAL,
    SHEET_DIAGNOSTICS,
    SHEET_FUND,
    SHEET_REGRESSION,
    SHEET_RISK,
    SHEET_ROLLING,
    SHEET_SUMMARY,
)

HEADER_FILL = "1F4E78"
SUBHEADER_FILL = "D9EAF7"
FORMULA_BLACK = "000000"
STATIC_GRAY = "666666"
THIN_GRAY = Side(style="thin", color="D9E2F3")
INT_FMT = "#,##0;[Red](#,##0);-"


def _section_title(ws, row: int, first_col: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    cell = ws.cell(row, first_col, text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="left")


def _add_guide(ws, start_row: int, first_col: int, last_col: int, title: str, rows: list[tuple[str, str]]) -> None:
    _section_title(ws, start_row, first_col, last_col, title)
    label_col = first_col
    text_start = first_col + 1
    for offset, (label, explanation) in enumerate(rows, start=1):
        row = start_row + offset
        ws.cell(row, label_col, label)
        ws.cell(row, label_col).font = Font(bold=True, color=STATIC_GRAY)
        ws.cell(row, label_col).fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
        if last_col > text_start:
            ws.merge_cells(start_row=row, start_column=text_start, end_row=row, end_column=last_col)
        text_cell = ws.cell(row, text_start, explanation)
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(first_col, last_col + 1):
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)
        ws.row_dimensions[row].height = 42


def _label_linest_output(ws, standard: bool) -> None:
    if standard:
        column_labels = ["Volatility", "Credit", "Currency", "Equity", "Alpha"]
        last_col = 6
    else:
        column_labels = [
            "Crisis x Volatility",
            "Crisis x Credit",
            "Crisis x Currency",
            "Crisis x Equity",
            "Volatility",
            "Credit",
            "Currency",
            "Equity",
            "Alpha",
        ]
        last_col = 10

    ws["A4"] = "LINEST row / term"
    for col, label in enumerate(column_labels, start=2):
        cell = ws.cell(4, col, label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws["A4"].font = Font(bold=True, color="FFFFFF", size=9)
    ws["A4"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A4"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 34

    row_labels = {
        6: "Coefficients",
        7: "Standard errors",
        8: "R-squared / SE(y)",
        9: "F-statistic / df",
        10: "SS regression / residual",
    }
    for row, label in row_labels.items():
        cell = ws.cell(row, 1, label)
        cell.font = Font(bold=True, color=STATIC_GRAY, size=9)
        cell.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 25)
    for col in range(2, last_col + 1):
        letter = ws.cell(4, col).column_letter
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 17)

    ws["A8"].comment = Comment(
        "LINEST row 3 uses only the first two result cells: R-squared and standard error of the y estimate. Remaining cells on this row are not model coefficients.",
        "Factor Risk Analysis",
    )
    ws["A9"].comment = Comment(
        "LINEST row 4 uses only the first two result cells: overall F-statistic and residual degrees of freedom.",
        "Factor Risk Analysis",
    )
    ws["A10"].comment = Comment(
        "LINEST row 5 uses only the first two result cells: regression sum of squares and residual sum of squares.",
        "Factor Risk Analysis",
    )


def _add_regression_guidance(ws, conditional: bool) -> None:
    rows = [
        ("Coefficient / beta", "Estimated sensitivity of monthly fund returns to that factor, holding the other factors constant. The sign gives direction; magnitude gives exposure."),
        ("Standard error", "Estimated uncertainty around the coefficient. A large standard error relative to the coefficient means the exposure is imprecisely estimated."),
        ("t-statistic", "Coefficient divided by its standard error. Larger absolute values provide stronger evidence that the coefficient differs from zero; roughly |t| near 2 is often associated with 5% significance in moderate/large samples."),
        ("p-value", "Probability, under the null hypothesis that the coefficient is zero, of observing a t-statistic at least this extreme. p < 0.05 is a common significance threshold, but statistical significance is not the same as economic importance."),
        ("R-squared", "Share of variation in monthly fund returns explained by the model in-sample. Higher is more explanatory, but not necessarily more predictive."),
        ("Adjusted R-squared", "R-squared adjusted for the number of predictors. Prefer this when comparing models with different numbers of factors."),
        ("F-statistic", "Joint test of whether all slope coefficients are zero. A larger F-statistic paired with a small model p-value indicates that the factor set is jointly informative."),
        ("Model p-value", "Significance level for the overall F-test. A small value, commonly below 0.05, suggests the regression has explanatory power as a whole."),
    ]
    if conditional:
        rows.extend(
            [
                ("Normal beta", "Estimated factor exposure in non-crisis months."),
                ("Crisis incremental beta", "Estimated change in that exposure when the crisis dummy equals 1. It is not the full crisis exposure by itself."),
                ("Effective crisis beta", "Normal beta plus crisis incremental beta; this is the estimated total exposure during crisis months."),
            ]
        )
    _add_guide(ws, 3, 12, 16, "Regression Interpretation Guide", rows)
    ws.column_dimensions["L"].width = 24
    for letter in ["M", "N", "O", "P"]:
        ws.column_dimensions[letter].width = 24


def _add_diagnostics_guidance(ws) -> None:
    rows = [
        ("Core correlation matrix", "Shows pairwise correlations among the fund and the four base factors. Fund-factor correlation is a simple two-variable relationship and should not be interpreted as a regression beta."),
        ("Factor correlations", "Large absolute correlations between predictors can indicate multicollinearity. As a practical flag, |correlation| above roughly 0.70-0.80 deserves attention because individual betas and p-values may become unstable."),
        ("Conditional matrix", "Adds crisis interaction terms. High correlations here are common because interaction variables are zero outside crisis months and may be based on a small crisis sample."),
        ("What to watch", "If coefficients change sharply when factors are added/removed, standard errors are large, or predictor correlations are very high, interpret individual exposures cautiously even when overall R-squared is strong."),
        ("Data quality", "Review the warning and dropped-month sections before interpreting results. Missing months reduce the effective sample and can alter factor estimates."),
    ]
    _add_guide(ws, 3, 12, 16, "How to Interpret Diagnostics", rows)
    ws.column_dimensions["L"].width = 25
    for letter in ["M", "N", "O", "P"]:
        ws.column_dimensions[letter].width = 24


def _add_rolling_guidance(ws) -> None:
    rows = [
        ("Rolling beta", "Each beta is re-estimated using only the selected trailing window ending in that month. Changes over time can reveal shifts in the fund's market exposures."),
        ("Rolling R-squared", "Percentage of return variation explained by the four factors within each trailing window. Rising R-squared suggests the fund is behaving more like the selected public-market factors; falling R-squared suggests more idiosyncratic behavior."),
        ("Window length", "Shorter windows react faster but are noisier; longer windows are more stable but can mask recent regime changes."),
        ("Blank early rows", "No rolling estimate is shown until a full window of observations is available."),
        ("Interpretation", "Focus on persistent changes rather than a single monthly jump. Large beta swings can reflect genuine exposure changes, changing correlations, or estimation noise."),
    ]
    _add_guide(ws, 3, 9, 13, "How to Interpret Rolling Analysis", rows)
    ws.column_dimensions["I"].width = 23
    for letter in ["J", "K", "L", "M"]:
        ws.column_dimensions[letter].width = 24


def _add_drawdown_duration(wb) -> None:
    ws_fund = wb[SHEET_FUND]
    first_row = 5
    last_row = ws_fund.max_row
    has_benchmark = any(ws_fund.cell(row, 3).value not in (None, "") for row in range(first_row, last_row + 1))

    # Helper columns keep the longest-drawdown metric auditable and fully formula-driven.
    ws_fund["N4"] = "Fund Underwater Months"
    ws_fund["O4"] = "Benchmark Underwater Months"
    for target in [ws_fund["N4"], ws_fund["O4"]]:
        target.font = copy(ws_fund["M4"].font)
        target.fill = copy(ws_fund["M4"].fill)
        target.alignment = copy(ws_fund["M4"].alignment)
        target.border = copy(ws_fund["M4"].border)
    ws_fund["N4"].comment = Comment(
        "Consecutive month-end observations below the prior high-water mark. The recovery month resets the count to zero; an ongoing drawdown is counted through the latest observation.",
        "Factor Risk Analysis",
    )

    for row in range(first_row, last_row + 1):
        if row == first_row:
            ws_fund.cell(row, 14, f"=IF(F{row}<0,1,0)")
            if has_benchmark:
                ws_fund.cell(row, 15, f"=IF(I{row}<0,1,0)")
        else:
            ws_fund.cell(row, 14, f"=IF(F{row}<0,N{row-1}+1,0)")
            if has_benchmark:
                ws_fund.cell(row, 15, f"=IF(I{row}<0,O{row-1}+1,0)")
        ws_fund.cell(row, 14).number_format = INT_FMT
        ws_fund.cell(row, 14).font = Font(color=FORMULA_BLACK)
        if has_benchmark:
            ws_fund.cell(row, 15).number_format = INT_FMT
            ws_fund.cell(row, 15).font = Font(color=FORMULA_BLACK)
    ws_fund.column_dimensions["N"].width = 22
    ws_fund.column_dimensions["O"].width = 28

    ws_risk = wb[SHEET_RISK]
    ws_risk["A24"] = "Longest drawdown (months)"
    ws_risk["B24"] = f"=MAX('Fund Returns'!$N${first_row}:$N${last_row})"
    ws_risk["B24"].number_format = INT_FMT
    ws_risk["B24"].font = Font(color=FORMULA_BLACK)
    if has_benchmark:
        ws_risk["C24"] = f"=MAX('Fund Returns'!$O${first_row}:$O${last_row})"
        ws_risk["C24"].number_format = INT_FMT
        ws_risk["C24"].font = Font(color=FORMULA_BLACK)
    ws_risk["A25"] = "Definition"
    ws_risk["B25"] = "Longest consecutive number of month-end observations below the prior peak; includes an ongoing unrecovered drawdown through the latest month."
    ws_risk.merge_cells("B25:F25")
    ws_risk["B25"].alignment = Alignment(wrap_text=True)
    ws_risk.row_dimensions[25].height = 32

    ws_summary = wb[SHEET_SUMMARY]
    ws_summary["A32"] = "Longest drawdown (months)"
    ws_summary["B32"] = f"='{SHEET_RISK}'!B24"
    ws_summary["B32"].number_format = INT_FMT
    ws_summary["B32"].font = Font(color=FORMULA_BLACK)
    if has_benchmark:
        ws_summary["C32"] = f"='{SHEET_RISK}'!C24"
        ws_summary["C32"].number_format = INT_FMT
        ws_summary["C32"].font = Font(color=FORMULA_BLACK)
    for col in range(1, 4):
        ws_summary.cell(32, col).border = Border(bottom=THIN_GRAY)


def _add_summary_f_stat_guidance(ws) -> None:
    start_row = max(45, ws.max_row + 2)
    rows = [
        ("What it tests", "The F-statistic tests whether the factor coefficients are jointly equal to zero. In other words, it asks whether the model explains more than an intercept-only model."),
        ("How to read it", "There is no universal 'good' F-statistic because its scale depends on sample size and number of predictors. Read it together with the model p-value."),
        ("Practical rule", "A large F-statistic with a small model p-value (commonly < 0.05) indicates that the factors are jointly statistically significant. It does not mean every individual factor is significant."),
        ("Standard vs conditional", "When comparing the standard and conditional models, also compare adjusted R-squared. The conditional model has more predictors, so a higher raw R-squared alone is not sufficient evidence that it is better."),
    ]
    _add_guide(ws, start_row, 1, 5, "How to Interpret the F-statistic", rows)


def _show_chart_axis_values(ws) -> None:
    charts = list(ws._charts)
    # Summary charts are added in a known order by workbook.py.
    y_formats = ["0.0", "0.0%", "0.0%", "0.00", "0.00", "0.0%"]
    for idx, chart in enumerate(charts):
        if getattr(chart, "x_axis", None) is not None:
            chart.x_axis.delete = False
            chart.x_axis.tickLblPos = "nextTo"
            chart.x_axis.majorTickMark = "out"
            if hasattr(chart.x_axis, "tickLblSkip"):
                chart.x_axis.tickLblSkip = 12
        if getattr(chart, "y_axis", None) is not None:
            chart.y_axis.delete = False
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.majorTickMark = "out"
            if idx < len(y_formats):
                chart.y_axis.numFmt = y_formats[idx]


def enhance_workbook(workbook_bytes: bytes) -> bytes:
    """Add interpretation aids, chart tick labels, and drawdown duration metrics."""

    wb = load_workbook(BytesIO(workbook_bytes))

    _label_linest_output(wb[SHEET_REGRESSION], standard=True)
    _label_linest_output(wb[SHEET_CONDITIONAL], standard=False)
    _add_regression_guidance(wb[SHEET_REGRESSION], conditional=False)
    _add_regression_guidance(wb[SHEET_CONDITIONAL], conditional=True)
    _add_diagnostics_guidance(wb[SHEET_DIAGNOSTICS])
    _add_rolling_guidance(wb[SHEET_ROLLING])
    _add_drawdown_duration(wb)
    _add_summary_f_stat_guidance(wb[SHEET_SUMMARY])
    _show_chart_axis_values(wb[SHEET_SUMMARY])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
