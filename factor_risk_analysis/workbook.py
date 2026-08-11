"""Excel workbook generation with live formulas."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    DEFAULT_CRISIS_THRESHOLD,
    DEFAULT_FACTORS,
    DEFAULT_ROLLING_WINDOW,
    FACTOR_NAMES,
    SHEET_CONDITIONAL,
    SHEET_DEFINITIONS,
    SHEET_DIAGNOSTICS,
    SHEET_FACTORS,
    SHEET_FUND,
    SHEET_REGRESSION,
    SHEET_RISK,
    SHEET_ROLLING,
    SHEET_SUMMARY,
    FactorDefinition,
)

DATE_FMT = "mmm-yy"
PCT_FMT = "0.0%;[Red](0.0%);-"
PCT2_FMT = "0.00%;[Red](0.00%);-"
NUM_FMT = "#,##0.00;[Red](#,##0.00);-"
INT_FMT = "#,##0;[Red](#,##0);-"
PVALUE_FMT = "0.000"
TEXT_FMT = "@"

HEADER_FILL = "1F4E78"
SUBHEADER_FILL = "D9EAF7"
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
LINK_GREEN = "008000"
STATIC_GRAY = "666666"
WARNING_FILL = "FCE4D6"
CONTROL_PURPLE = "7030A0"

THIN_GRAY = Side(style="thin", color="D9E2F3")
MEDIUM_BLUE = Side(style="medium", color=HEADER_FILL)


@dataclass(frozen=True)
class WorkbookSettings:
    """Inputs and metadata written to the generated workbook."""

    fund_name: str
    input_mode: str
    fund_source: str
    requested_start: str
    requested_end: str
    risk_free_rate_annual: float
    rolling_window: int = DEFAULT_ROLLING_WINDOW
    crisis_threshold: float = DEFAULT_CRISIS_THRESHOLD
    factors: Sequence[FactorDefinition] = DEFAULT_FACTORS
    benchmark_name: str | None = None
    return_scale_method: str | None = None
    data_warnings: Sequence[str] = field(default_factory=tuple)
    dropped_rows: int = 0
    factor_metadata: Sequence[Mapping[str, object]] = field(default_factory=tuple)


def qsheet(sheet_name: str) -> str:
    return f"'{sheet_name.replace("'", "''")}'"


def xl_range(sheet_name: str, col: int, first_row: int, last_row: int) -> str:
    letter = get_column_letter(col)
    return f"{qsheet(sheet_name)}!${letter}${first_row}:${letter}${last_row}"


def xl_block(sheet_name: str, first_col: int, first_row: int, last_col: int, last_row: int) -> str:
    first_letter = get_column_letter(first_col)
    last_letter = get_column_letter(last_col)
    return f"{qsheet(sheet_name)}!${first_letter}${first_row}:${last_letter}${last_row}"


def style_title(ws: Worksheet, title: str, last_col: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 24


def style_section(ws: Worksheet, row: int, title: str, first_col: int = 1, last_col: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    cell = ws.cell(row, first_col, title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="left")


def style_header_row(ws: Worksheet, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=MEDIUM_BLUE)


def set_widths(ws: Worksheet, widths: Mapping[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def format_percent(cell, decimals: int = 1) -> None:
    cell.number_format = PCT_FMT if decimals == 1 else PCT2_FMT


def formula_cell(cell) -> None:
    cell.font = Font(color=FORMULA_BLACK)


def input_cell(cell) -> None:
    cell.font = Font(color=INPUT_BLUE)


def imported_cell(cell) -> None:
    cell.font = Font(color=LINK_GREEN)


def static_cell(cell) -> None:
    cell.font = Font(color=STATIC_GRAY)


def apply_grid(ws: Worksheet, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    for row in range(first_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            ws.cell(row, col).border = Border(bottom=THIN_GRAY)


def add_note(cell, text: str) -> None:
    cell.comment = Comment(text, "Factor Risk Analysis")


def create_workbook(data: pd.DataFrame, settings: WorkbookSettings) -> bytes:
    """Create an auditable Microsoft 365 workbook and return its bytes."""

    if data.empty:
        raise ValueError("Cannot create workbook from an empty data set.")

    missing = [name for name in FACTOR_NAMES if name not in data.columns]
    if missing:
        raise ValueError(f"Missing required factor columns: {missing}")
    if "Fund" not in data.columns:
        raise ValueError("Missing Fund return column.")
    if "VIX Level" not in data.columns:
        raise ValueError("Missing VIX Level column.")

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws_summary = wb.create_sheet(SHEET_SUMMARY)
    ws_fund = wb.create_sheet(SHEET_FUND)
    ws_factors = wb.create_sheet(SHEET_FACTORS)
    ws_reg = wb.create_sheet(SHEET_REGRESSION)
    ws_cond = wb.create_sheet(SHEET_CONDITIONAL)
    ws_risk = wb.create_sheet(SHEET_RISK)
    ws_diag = wb.create_sheet(SHEET_DIAGNOSTICS)
    ws_roll = wb.create_sheet(SHEET_ROLLING)
    ws_defs = wb.create_sheet(SHEET_DEFINITIONS)

    has_benchmark = "Benchmark" in data.columns
    data = data.copy().sort_index()
    n_obs = len(data)
    first_data_row = 5
    last_data_row = first_data_row + n_obs - 1

    _write_fund_returns_sheet(ws_fund, data, settings, first_data_row, last_data_row, has_benchmark)
    _write_factor_data_sheet(ws_factors, data, settings, first_data_row, last_data_row)
    _write_regression_sheet(ws_reg, data, settings, first_data_row, last_data_row)
    _write_conditional_regression_sheet(ws_cond, data, settings, first_data_row, last_data_row)
    _write_risk_statistics_sheet(ws_risk, data, settings, first_data_row, last_data_row, has_benchmark)
    _write_diagnostics_sheet(ws_diag, data, settings, first_data_row, last_data_row)
    _write_rolling_analysis_sheet(ws_roll, data, settings, first_data_row, last_data_row)
    _write_definitions_sheet(ws_defs, data, settings)
    _write_summary_sheet(ws_summary, data, settings, first_data_row, last_data_row, has_benchmark)
    _add_summary_charts(wb, data, first_data_row, last_data_row, has_benchmark)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5" if ws.title in {SHEET_FUND, SHEET_FACTORS, SHEET_ROLLING} else None

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_fund_returns_sheet(
    ws: Worksheet,
    data: pd.DataFrame,
    settings: WorkbookSettings,
    first_row: int,
    last_row: int,
    has_benchmark: bool,
) -> None:
    style_title(ws, "Fund and Benchmark Returns", 13)
    ws["A3"] = "Base index"
    ws["D3"] = 100
    ws["G3"] = 100 if has_benchmark else None
    static_cell(ws["A3"])
    input_cell(ws["D3"])
    if has_benchmark:
        input_cell(ws["G3"])
    ws["D3"].number_format = NUM_FMT
    ws["G3"].number_format = NUM_FMT

    headers = [
        "Date",
        "Fund Return",
        "Benchmark Return",
        "Fund Wealth Index",
        "Fund Peak",
        "Fund Drawdown",
        "Benchmark Wealth Index",
        "Benchmark Peak",
        "Benchmark Drawdown",
        "Fund Rolling 12M Return",
        "Benchmark Rolling 12M Return",
        "Fund Downside Sq",
        "Benchmark Downside Sq",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)
    style_header_row(ws, 4, 1, len(headers))

    add_note(ws["B4"], f"Fund source: {settings.fund_source}")
    if has_benchmark:
        add_note(ws["C4"], "Benchmark monthly total return calculated from Yahoo Finance adjusted close.")

    for i, (idx, row) in enumerate(data.iterrows(), start=0):
        excel_row = first_row + i
        ws.cell(excel_row, 1, idx.to_pydatetime())
        ws.cell(excel_row, 1).number_format = DATE_FMT
        imported_cell(ws.cell(excel_row, 1))

        ws.cell(excel_row, 2, float(row["Fund"]))
        format_percent(ws.cell(excel_row, 2), 2)
        imported_cell(ws.cell(excel_row, 2))

        if has_benchmark:
            ws.cell(excel_row, 3, float(row["Benchmark"]))
            format_percent(ws.cell(excel_row, 3), 2)
            imported_cell(ws.cell(excel_row, 3))

        # Wealth, peak, drawdown.
        if i == 0:
            ws.cell(excel_row, 4, f"=$D$3*(1+B{excel_row})")
        else:
            ws.cell(excel_row, 4, f"=D{excel_row - 1}*(1+B{excel_row})")
        ws.cell(excel_row, 5, f"=MAX($D$3:D{excel_row})")
        ws.cell(excel_row, 6, f"=D{excel_row}/E{excel_row}-1")
        for col in (4, 5):
            ws.cell(excel_row, col).number_format = NUM_FMT
            formula_cell(ws.cell(excel_row, col))
        format_percent(ws.cell(excel_row, 6), 1)
        formula_cell(ws.cell(excel_row, 6))

        if has_benchmark:
            if i == 0:
                ws.cell(excel_row, 7, f"=$G$3*(1+C{excel_row})")
            else:
                ws.cell(excel_row, 7, f"=G{excel_row - 1}*(1+C{excel_row})")
            ws.cell(excel_row, 8, f"=MAX($G$3:G{excel_row})")
            ws.cell(excel_row, 9, f"=G{excel_row}/H{excel_row}-1")
            for col in (7, 8):
                ws.cell(excel_row, col).number_format = NUM_FMT
                formula_cell(ws.cell(excel_row, col))
            format_percent(ws.cell(excel_row, 9), 1)
            formula_cell(ws.cell(excel_row, 9))

        # Rolling 12-month compounded returns.
        if i >= 11:
            if i == 11:
                ws.cell(excel_row, 10, f"=D{excel_row}/$D$3-1")
                if has_benchmark:
                    ws.cell(excel_row, 11, f"=G{excel_row}/$G$3-1")
            else:
                ws.cell(excel_row, 10, f"=D{excel_row}/D{excel_row - 12}-1")
                if has_benchmark:
                    ws.cell(excel_row, 11, f"=G{excel_row}/G{excel_row - 12}-1")
        else:
            ws.cell(excel_row, 10, "")
            if has_benchmark:
                ws.cell(excel_row, 11, "")
        format_percent(ws.cell(excel_row, 10), 1)
        formula_cell(ws.cell(excel_row, 10))
        if has_benchmark:
            format_percent(ws.cell(excel_row, 11), 1)
            formula_cell(ws.cell(excel_row, 11))

        ws.cell(excel_row, 12, f"=MIN(B{excel_row},0)^2")
        ws.cell(excel_row, 12).number_format = "0.000000"
        formula_cell(ws.cell(excel_row, 12))
        if has_benchmark:
            ws.cell(excel_row, 13, f"=MIN(C{excel_row},0)^2")
            ws.cell(excel_row, 13).number_format = "0.000000"
            formula_cell(ws.cell(excel_row, 13))

    set_widths(
        ws,
        {
            "A": 12,
            "B": 14,
            "C": 18,
            "D": 18,
            "E": 14,
            "F": 15,
            "G": 22,
            "H": 18,
            "I": 20,
            "J": 22,
            "K": 26,
            "L": 18,
            "M": 22,
        },
    )
    apply_grid(ws, first_row, last_row, 1, 13)


def _write_factor_data_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings, first_row: int, last_row: int) -> None:
    style_title(ws, "Factor Data", 11)
    ws["A3"] = "Crisis threshold"
    ws["B3"] = settings.crisis_threshold
    ws["B3"].number_format = NUM_FMT
    input_cell(ws["B3"])
    add_note(ws["B3"], "Crisis month if month-end VIX level is above this threshold.")

    headers = [
        "Date",
        "Equity Return",
        "Currency Return",
        "Credit Return",
        "Volatility Return",
        "Crisis x Equity",
        "Crisis x Currency",
        "Crisis x Credit",
        "Crisis x Volatility",
        "VIX Level",
        "Crisis Dummy",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)
    style_header_row(ws, 4, 1, len(headers))

    factor_by_name = {factor.name: factor for factor in settings.factors}
    for col, name in zip(range(2, 6), FACTOR_NAMES, strict=True):
        factor = factor_by_name[name]
        add_note(
            ws.cell(4, col),
            f"Source: Yahoo Finance via yfinance. Ticker: {factor.ticker}. Method: {factor.method}.",
        )
    add_note(ws["J4"], "Source: Yahoo Finance via yfinance. Month-end VIX level used for crisis classification.")

    for i, (idx, row) in enumerate(data.iterrows(), start=0):
        excel_row = first_row + i
        ws.cell(excel_row, 1, idx.to_pydatetime())
        ws.cell(excel_row, 1).number_format = DATE_FMT
        imported_cell(ws.cell(excel_row, 1))

        for offset, factor_name in enumerate(FACTOR_NAMES, start=2):
            ws.cell(excel_row, offset, float(row[factor_name]))
            format_percent(ws.cell(excel_row, offset), 2)
            imported_cell(ws.cell(excel_row, offset))

        # Crisis interaction columns are formula-driven.
        ws.cell(excel_row, 10, float(row["VIX Level"]))
        ws.cell(excel_row, 10).number_format = NUM_FMT
        imported_cell(ws.cell(excel_row, 10))

        ws.cell(excel_row, 11, f"=--(J{excel_row}>$B$3)")
        ws.cell(excel_row, 11).number_format = INT_FMT
        formula_cell(ws.cell(excel_row, 11))

        for col, base_col in zip(range(6, 10), range(2, 6), strict=True):
            ws.cell(excel_row, col, f"=K{excel_row}*{get_column_letter(base_col)}{excel_row}")
            format_percent(ws.cell(excel_row, col), 2)
            formula_cell(ws.cell(excel_row, col))

    set_widths(
        ws,
        {
            "A": 12,
            "B": 16,
            "C": 18,
            "D": 15,
            "E": 18,
            "F": 18,
            "G": 20,
            "H": 18,
            "I": 22,
            "J": 12,
            "K": 14,
        },
    )
    apply_grid(ws, first_row, last_row, 1, 11)


def _write_regression_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings, first_row: int, last_row: int) -> None:
    style_title(ws, "Standard Four-Factor Regression", 9)
    ws["A3"] = "Model"
    ws["B3"] = "Fund return = alpha + equity + currency + credit + volatility + error"
    static_cell(ws["A3"])
    static_cell(ws["B3"])

    # LINEST output block.
    style_section(ws, 5, "LINEST Output", 1, 9)
    linest_ref = "B6:F10"
    formula = f"=LINEST({xl_range(SHEET_FUND, 2, first_row, last_row)},{xl_block(SHEET_FACTORS, 2, first_row, 5, last_row)},TRUE,TRUE)"
    ws["B6"] = ArrayFormula(linest_ref, formula)
    add_note(ws["B6"], "Excel LINEST output. Coefficients are returned in reverse order of the known_x columns, with intercept last.")
    for row in range(6, 11):
        for col in range(2, 7):
            ws.cell(row, col).number_format = NUM_FMT
            formula_cell(ws.cell(row, col))

    # Coefficient table.
    style_section(ws, 12, "Coefficient Summary", 1, 9)
    headers = ["Term", "Coefficient", "Standard Error", "t-stat", "p-value"]
    for col, header in enumerate(headers, start=1):
        ws.cell(13, col, header)
    style_header_row(ws, 13, 1, len(headers))

    terms = ["Alpha", "Equity", "Currency", "Credit", "Volatility"]
    line_cols = {"Alpha": 5, "Equity": 4, "Currency": 3, "Credit": 2, "Volatility": 1}
    for i, term in enumerate(terms, start=14):
        ws.cell(i, 1, term)
        ws.cell(i, 2, f"=INDEX($B$6:$F$10,1,{line_cols[term]})")
        ws.cell(i, 3, f"=INDEX($B$6:$F$10,2,{line_cols[term]})")
        ws.cell(i, 4, f"=IFERROR(B{i}/C{i},\"\")")
        ws.cell(i, 5, f"=IFERROR(T.DIST.2T(ABS(D{i}),$C$9),\"\")")
        for col in range(2, 5):
            ws.cell(i, col).number_format = NUM_FMT
            formula_cell(ws.cell(i, col))
        ws.cell(i, 5).number_format = PVALUE_FMT
        formula_cell(ws.cell(i, 5))

    # Model statistics.
    style_section(ws, 21, "Model Statistics", 1, 9)
    stats = [
        ("Observations", f"=COUNT({xl_range(SHEET_FUND, 2, first_row, last_row)})", INT_FMT),
        ("Predictors", "=4", INT_FMT),
        ("R-squared", "=INDEX($B$6:$F$10,3,1)", PCT2_FMT),
        ("Adjusted R-squared", "=1-(1-B24)*(B22-1)/(B22-B23-1)", PCT2_FMT),
        ("F-statistic", "=INDEX($B$6:$F$10,4,1)", NUM_FMT),
        ("Model p-value", "=F.DIST.RT(B26,B23,B22-B23-1)", PVALUE_FMT),
    ]
    for i, (label, formula_text, number_format) in enumerate(stats, start=22):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula_text)
        ws.cell(i, 2).number_format = number_format
        formula_cell(ws.cell(i, 2))

    # Actual vs predicted.
    style_section(ws, 30, "Actual vs. Predicted", 1, 9)
    avp_headers = ["Date", "Actual Fund Return", "Predicted Return", "Residual"]
    for col, header in enumerate(avp_headers, start=1):
        ws.cell(31, col, header)
    style_header_row(ws, 31, 1, len(avp_headers))

    for i in range(len(data)):
        src_row = first_row + i
        out_row = 32 + i
        ws.cell(out_row, 1, f"={qsheet(SHEET_FUND)}!A{src_row}")
        ws.cell(out_row, 2, f"={qsheet(SHEET_FUND)}!B{src_row}")
        ws.cell(
            out_row,
            3,
            f"=$B$14+$B$15*{qsheet(SHEET_FACTORS)}!B{src_row}+$B$16*{qsheet(SHEET_FACTORS)}!C{src_row}"
            f"+$B$17*{qsheet(SHEET_FACTORS)}!D{src_row}+$B$18*{qsheet(SHEET_FACTORS)}!E{src_row}",
        )
        ws.cell(out_row, 4, f"=B{out_row}-C{out_row}")
        ws.cell(out_row, 1).number_format = DATE_FMT
        for col in (2, 3, 4):
            format_percent(ws.cell(out_row, col), 2)
            formula_cell(ws.cell(out_row, col))

    set_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 14, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12})
    apply_grid(ws, 14, 18, 1, 5)
    apply_grid(ws, 22, 27, 1, 2)
    apply_grid(ws, 32, 31 + len(data), 1, 4)


def _write_conditional_regression_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings, first_row: int, last_row: int) -> None:
    style_title(ws, "Conditional Four-Factor Regression", 10)
    ws["A3"] = "Model"
    ws["B3"] = "Fund return = alpha + normal factor betas + crisis dummy x factor incremental betas + error"
    static_cell(ws["A3"])
    static_cell(ws["B3"])

    style_section(ws, 5, "LINEST Output", 1, 10)
    linest_ref = "B6:J10"
    formula = f"=LINEST({xl_range(SHEET_FUND, 2, first_row, last_row)},{xl_block(SHEET_FACTORS, 2, first_row, 9, last_row)},TRUE,TRUE)"
    ws["B6"] = ArrayFormula(linest_ref, formula)
    add_note(ws["B6"], "Excel LINEST output for normal factor returns and crisis interaction columns.")
    for row in range(6, 11):
        for col in range(2, 11):
            ws.cell(row, col).number_format = NUM_FMT
            formula_cell(ws.cell(row, col))

    style_section(ws, 12, "Coefficient Summary", 1, 10)
    headers = [
        "Term",
        "Normal Beta",
        "Normal SE",
        "Normal p-value",
        "Crisis Incremental Beta",
        "Crisis Incremental SE",
        "Crisis Incremental p-value",
        "Effective Crisis Beta",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(13, col, header)
    style_header_row(ws, 13, 1, len(headers))

    # Column positions within the 9-column LINEST output block B:J.
    normal_cols = {"Alpha": 9, "Equity": 8, "Currency": 7, "Credit": 6, "Volatility": 5}
    crisis_cols = {"Equity": 4, "Currency": 3, "Credit": 2, "Volatility": 1}
    terms = ["Alpha", "Equity", "Currency", "Credit", "Volatility"]

    for i, term in enumerate(terms, start=14):
        ws.cell(i, 1, term)
        ws.cell(i, 2, f"=INDEX($B$6:$J$10,1,{normal_cols[term]})")
        ws.cell(i, 3, f"=INDEX($B$6:$J$10,2,{normal_cols[term]})")
        ws.cell(i, 4, f"=IFERROR(T.DIST.2T(ABS(B{i}/C{i}),$C$9),\"\")")
        if term == "Alpha":
            ws.cell(i, 5, "")
            ws.cell(i, 6, "")
            ws.cell(i, 7, "")
            ws.cell(i, 8, f"=B{i}")
        else:
            ws.cell(i, 5, f"=INDEX($B$6:$J$10,1,{crisis_cols[term]})")
            ws.cell(i, 6, f"=INDEX($B$6:$J$10,2,{crisis_cols[term]})")
            ws.cell(i, 7, f"=IFERROR(T.DIST.2T(ABS(E{i}/F{i}),$C$9),\"\")")
            ws.cell(i, 8, f"=B{i}+E{i}")
        for col in (2, 3, 5, 6, 8):
            ws.cell(i, col).number_format = NUM_FMT
            formula_cell(ws.cell(i, col))
        for col in (4, 7):
            ws.cell(i, col).number_format = PVALUE_FMT
            formula_cell(ws.cell(i, col))

    style_section(ws, 21, "Model Statistics", 1, 10)
    stats = [
        ("Observations", f"=COUNT({xl_range(SHEET_FUND, 2, first_row, last_row)})", INT_FMT),
        ("Predictors", "=8", INT_FMT),
        ("R-squared", "=INDEX($B$6:$J$10,3,1)", PCT2_FMT),
        ("Adjusted R-squared", "=1-(1-B24)*(B22-1)/(B22-B23-1)", PCT2_FMT),
        ("F-statistic", "=INDEX($B$6:$J$10,4,1)", NUM_FMT),
        ("Model p-value", "=F.DIST.RT(B26,B23,B22-B23-1)", PVALUE_FMT),
    ]
    for i, (label, formula_text, number_format) in enumerate(stats, start=22):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula_text)
        ws.cell(i, 2).number_format = number_format
        formula_cell(ws.cell(i, 2))

    style_section(ws, 30, "Actual vs. Conditional Predicted", 1, 10)
    headers = ["Date", "Actual Fund Return", "Predicted Return", "Residual"]
    for col, header in enumerate(headers, start=1):
        ws.cell(31, col, header)
    style_header_row(ws, 31, 1, len(headers))

    for i in range(len(data)):
        src_row = first_row + i
        out_row = 32 + i
        ws.cell(out_row, 1, f"={qsheet(SHEET_FUND)}!A{src_row}")
        ws.cell(out_row, 2, f"={qsheet(SHEET_FUND)}!B{src_row}")
        ws.cell(
            out_row,
            3,
            f"=$B$14+$B$15*{qsheet(SHEET_FACTORS)}!B{src_row}+$B$16*{qsheet(SHEET_FACTORS)}!C{src_row}"
            f"+$B$17*{qsheet(SHEET_FACTORS)}!D{src_row}+$B$18*{qsheet(SHEET_FACTORS)}!E{src_row}"
            f"+$E$15*{qsheet(SHEET_FACTORS)}!F{src_row}+$E$16*{qsheet(SHEET_FACTORS)}!G{src_row}"
            f"+$E$17*{qsheet(SHEET_FACTORS)}!H{src_row}+$E$18*{qsheet(SHEET_FACTORS)}!I{src_row}",
        )
        ws.cell(out_row, 4, f"=B{out_row}-C{out_row}")
        ws.cell(out_row, 1).number_format = DATE_FMT
        for col in (2, 3, 4):
            format_percent(ws.cell(out_row, col), 2)
            formula_cell(ws.cell(out_row, col))

    set_widths(
        ws,
        {
            "A": 18,
            "B": 16,
            "C": 14,
            "D": 16,
            "E": 22,
            "F": 20,
            "G": 24,
            "H": 20,
            "I": 12,
            "J": 12,
        },
    )
    apply_grid(ws, 14, 18, 1, 8)
    apply_grid(ws, 22, 27, 1, 2)
    apply_grid(ws, 32, 31 + len(data), 1, 4)


def _write_risk_statistics_sheet(
    ws: Worksheet,
    data: pd.DataFrame,
    settings: WorkbookSettings,
    first_row: int,
    last_row: int,
    has_benchmark: bool,
) -> None:
    style_title(ws, "Risk and Performance Statistics", 6)
    style_section(ws, 3, "Assumptions", 1, 6)
    ws["A4"] = "Annual risk-free rate"
    ws["B4"] = settings.risk_free_rate_annual
    ws["B4"].number_format = PCT2_FMT
    input_cell(ws["B4"])
    add_note(ws["B4"], "User-selected annual risk-free rate used for Sharpe and Sortino ratios.")
    ws["A5"] = "Monthly risk-free rate"
    ws["B5"] = "=(1+B4)^(1/12)-1"
    ws["B5"].number_format = PCT2_FMT
    formula_cell(ws["B5"])

    style_section(ws, 7, "Return and Risk Metrics", 1, 6)
    ws["A8"] = "Metric"
    ws["B8"] = settings.fund_name or "Fund"
    ws["C8"] = settings.benchmark_name or "Benchmark"
    style_header_row(ws, 8, 1, 3)

    fund_ret = xl_range(SHEET_FUND, 2, first_row, last_row)
    fund_dd = xl_range(SHEET_FUND, 6, first_row, last_row)
    fund_roll = xl_range(SHEET_FUND, 10, first_row, last_row)
    fund_downside_sq = xl_range(SHEET_FUND, 12, first_row, last_row)
    bench_ret = xl_range(SHEET_FUND, 3, first_row, last_row)
    bench_dd = xl_range(SHEET_FUND, 9, first_row, last_row)
    bench_roll = xl_range(SHEET_FUND, 11, first_row, last_row)
    bench_downside_sq = xl_range(SHEET_FUND, 13, first_row, last_row)

    metrics = [
        ("Complete monthly observations", f"=COUNT({fund_ret})", f"=IFERROR(COUNT({bench_ret}),\"\")", INT_FMT),
        ("First aligned month", f"=MIN({xl_range(SHEET_FUND, 1, first_row, last_row)})", f"=MIN({xl_range(SHEET_FUND, 1, first_row, last_row)})", DATE_FMT),
        ("Last aligned month", f"=MAX({xl_range(SHEET_FUND, 1, first_row, last_row)})", f"=MAX({xl_range(SHEET_FUND, 1, first_row, last_row)})", DATE_FMT),
        ("Monthly arithmetic return", f"=AVERAGE({fund_ret})", f"=IFERROR(AVERAGE({bench_ret}),\"\")", PCT2_FMT),
        ("Annualized arithmetic return", "=B12*12", "=IFERROR(C12*12,\"\")", PCT2_FMT),
        ("Monthly standard deviation", f"=STDEV.S({fund_ret})", f"=IFERROR(STDEV.S({bench_ret}),\"\")", PCT2_FMT),
        ("Annualized standard deviation", "=B14*SQRT(12)", "=IFERROR(C14*SQRT(12),\"\")", PCT2_FMT),
        ("Monthly downside deviation below 0", f"=SQRT(AVERAGE({fund_downside_sq}))", f"=IFERROR(SQRT(AVERAGE({bench_downside_sq})),\"\")", PCT2_FMT),
        ("Annualized downside deviation", "=B16*SQRT(12)", "=IFERROR(C16*SQRT(12),\"\")", PCT2_FMT),
        ("Sharpe ratio", "=IFERROR((B13-$B$4)/B15,\"\")", "=IFERROR((C13-$B$4)/C15,\"\")", NUM_FMT),
        ("Sortino ratio", "=IFERROR((B13-$B$4)/B17,\"\")", "=IFERROR((C13-$B$4)/C17,\"\")", NUM_FMT),
        ("Maximum drawdown", f"=MIN({fund_dd})", f"=IFERROR(MIN({bench_dd}),\"\")", PCT2_FMT),
        ("Annual 95% historical VaR", f"=IF(COUNT({fund_roll})>0,-PERCENTILE.INC({fund_roll},0.05),\"\")", f"=IFERROR(IF(COUNT({bench_roll})>0,-PERCENTILE.INC({bench_roll},0.05),\"\"),\"\")", PCT2_FMT),
        ("Annual 95% historical CVaR", f"=IF(COUNT({fund_roll})>0,-AVERAGEIF({fund_roll},\"<=\"&PERCENTILE.INC({fund_roll},0.05),{fund_roll}),\"\")", f"=IFERROR(IF(COUNT({bench_roll})>0,-AVERAGEIF({bench_roll},\"<=\"&PERCENTILE.INC({bench_roll},0.05),{bench_roll}),\"\"),\"\")", PCT2_FMT),
        ("Rolling 12M observations", f"=COUNT({fund_roll})", f"=IFERROR(COUNT({bench_roll}),\"\")", INT_FMT),
    ]

    for row_idx, (label, fund_formula, bench_formula, number_format) in enumerate(metrics, start=9):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, fund_formula)
        ws.cell(row_idx, 2).number_format = number_format
        formula_cell(ws.cell(row_idx, 2))
        if has_benchmark:
            ws.cell(row_idx, 3, bench_formula)
            ws.cell(row_idx, 3).number_format = number_format
            formula_cell(ws.cell(row_idx, 3))
        else:
            ws.cell(row_idx, 3, "")

    style_section(ws, 27, "Benchmark Analytics", 1, 6)
    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, start=1):
        ws.cell(28, col, header)
    style_header_row(ws, 28, 1, 2)

    bench_metrics = [
        ("Fund / benchmark correlation", f"=IFERROR(CORREL({fund_ret},{bench_ret}),\"\")", NUM_FMT),
        ("Fund beta to benchmark", f"=IFERROR(SLOPE({fund_ret},{bench_ret}),\"\")", NUM_FMT),
        ("Upside capture", f"=IFERROR(AVERAGEIFS({fund_ret},{bench_ret},\">0\")/AVERAGEIFS({bench_ret},{bench_ret},\">0\"),\"\")", PCT2_FMT),
        ("Downside capture", f"=IFERROR(AVERAGEIFS({fund_ret},{bench_ret},\"<0\")/AVERAGEIFS({bench_ret},{bench_ret},\"<0\"),\"\")", PCT2_FMT),
    ]
    for row_idx, (label, formula_text, number_format) in enumerate(bench_metrics, start=29):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, formula_text if has_benchmark else "")
        ws.cell(row_idx, 2).number_format = number_format
        formula_cell(ws.cell(row_idx, 2))

    set_widths(ws, {"A": 36, "B": 20, "C": 20, "D": 14, "E": 14, "F": 14})
    apply_grid(ws, 9, 23, 1, 3)
    apply_grid(ws, 29, 32, 1, 2)


def _write_diagnostics_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings, first_row: int, last_row: int) -> None:
    style_title(ws, "Diagnostics", 10)
    style_section(ws, 3, "Core Factor Correlation Matrix", 1, 8)

    variables = ["Fund", "Equity", "Currency", "Credit", "Volatility"]
    ranges = {
        "Fund": xl_range(SHEET_FUND, 2, first_row, last_row),
        "Equity": xl_range(SHEET_FACTORS, 2, first_row, last_row),
        "Currency": xl_range(SHEET_FACTORS, 3, first_row, last_row),
        "Credit": xl_range(SHEET_FACTORS, 4, first_row, last_row),
        "Volatility": xl_range(SHEET_FACTORS, 5, first_row, last_row),
    }
    for j, name in enumerate(variables, start=2):
        ws.cell(4, j, name)
    for i, name in enumerate(variables, start=5):
        ws.cell(i, 1, name)
        for j, name2 in enumerate(variables, start=2):
            ws.cell(i, j, f"=IFERROR(CORREL({ranges[name]},{ranges[name2]}),\"\")")
            ws.cell(i, j).number_format = NUM_FMT
            formula_cell(ws.cell(i, j))
    style_header_row(ws, 4, 1, len(variables) + 1)
    apply_grid(ws, 5, 5 + len(variables) - 1, 1, len(variables) + 1)

    style_section(ws, 12, "Conditional Predictor Correlation Matrix", 1, 10)
    cond_vars = ["Equity", "Currency", "Credit", "Volatility", "Crisis x Equity", "Crisis x Currency", "Crisis x Credit", "Crisis x Volatility"]
    cond_ranges = {
        "Equity": xl_range(SHEET_FACTORS, 2, first_row, last_row),
        "Currency": xl_range(SHEET_FACTORS, 3, first_row, last_row),
        "Credit": xl_range(SHEET_FACTORS, 4, first_row, last_row),
        "Volatility": xl_range(SHEET_FACTORS, 5, first_row, last_row),
        "Crisis x Equity": xl_range(SHEET_FACTORS, 6, first_row, last_row),
        "Crisis x Currency": xl_range(SHEET_FACTORS, 7, first_row, last_row),
        "Crisis x Credit": xl_range(SHEET_FACTORS, 8, first_row, last_row),
        "Crisis x Volatility": xl_range(SHEET_FACTORS, 9, first_row, last_row),
    }
    for j, name in enumerate(cond_vars, start=2):
        ws.cell(13, j, name)
        ws.cell(13, j).alignment = Alignment(text_rotation=45)
    for i, name in enumerate(cond_vars, start=14):
        ws.cell(i, 1, name)
        for j, name2 in enumerate(cond_vars, start=2):
            ws.cell(i, j, f"=IFERROR(CORREL({cond_ranges[name]},{cond_ranges[name2]}),\"\")")
            ws.cell(i, j).number_format = NUM_FMT
            formula_cell(ws.cell(i, j))
    style_header_row(ws, 13, 1, len(cond_vars) + 1)
    apply_grid(ws, 14, 14 + len(cond_vars) - 1, 1, len(cond_vars) + 1)

    style_section(ws, 26, "Data Quality and Warnings", 1, 10)
    ws["A27"] = "Complete observations"
    ws["B27"] = f"=COUNT({xl_range(SHEET_FUND, 2, first_row, last_row)})"
    ws["B27"].number_format = INT_FMT
    formula_cell(ws["B27"])
    ws["A28"] = "Dropped observations"
    ws["B28"] = settings.dropped_rows
    ws["B28"].number_format = INT_FMT
    input_cell(ws["B28"])
    ws["A29"] = "Crisis months"
    ws["B29"] = f"=SUM({xl_range(SHEET_FACTORS, 11, first_row, last_row)})"
    ws["B29"].number_format = INT_FMT
    formula_cell(ws["B29"])
    ws["A30"] = "Warnings"
    if settings.data_warnings:
        for idx, warning in enumerate(settings.data_warnings, start=31):
            ws.cell(idx, 1, warning)
            ws.cell(idx, 1).fill = PatternFill("solid", fgColor=WARNING_FILL)
    else:
        ws["A31"] = "No data-quality warnings."

    set_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 18, "G": 18, "H": 18, "I": 20, "J": 18})


def _write_rolling_analysis_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings, first_row: int, last_row: int) -> None:
    style_title(ws, "Rolling Regression Analysis", 8)
    ws["A3"] = "Rolling window (months)"
    ws["B3"] = settings.rolling_window
    ws["B3"].number_format = INT_FMT
    input_cell(ws["B3"])

    headers = ["Date", "Equity Beta", "Currency Beta", "Credit Beta", "Volatility Beta", "Rolling R-squared", "Observations"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)
    style_header_row(ws, 4, 1, len(headers))

    window = int(settings.rolling_window)
    for i, idx in enumerate(data.index, start=0):
        out_row = first_row + i
        data_row = first_row + i
        ws.cell(out_row, 1, f"={qsheet(SHEET_FUND)}!A{data_row}")
        ws.cell(out_row, 1).number_format = DATE_FMT
        formula_cell(ws.cell(out_row, 1))
        if i + 1 < window:
            for col in range(2, 8):
                ws.cell(out_row, col, "")
            continue

        start_src_row = data_row - window + 1
        end_src_row = data_row
        y_range = xl_range(SHEET_FUND, 2, start_src_row, end_src_row)
        x_range = xl_block(SHEET_FACTORS, 2, start_src_row, 5, end_src_row)
        linest = f"LINEST({y_range},{x_range},TRUE,TRUE)"
        coef_cols = [4, 3, 2, 1]
        for out_col, linest_col in zip(range(2, 6), coef_cols, strict=True):
            ws.cell(out_row, out_col, f"=IFERROR(INDEX({linest},1,{linest_col}),\"\")")
            ws.cell(out_row, out_col).number_format = NUM_FMT
            formula_cell(ws.cell(out_row, out_col))
        ws.cell(out_row, 6, f"=IFERROR(INDEX({linest},3,1),\"\")")
        ws.cell(out_row, 6).number_format = PCT2_FMT
        formula_cell(ws.cell(out_row, 6))
        ws.cell(out_row, 7, f"=COUNT({y_range})")
        ws.cell(out_row, 7).number_format = INT_FMT
        formula_cell(ws.cell(out_row, 7))

    set_widths(ws, {"A": 12, "B": 14, "C": 15, "D": 14, "E": 16, "F": 18, "G": 14})
    apply_grid(ws, first_row, last_row, 1, 7)


def _write_definitions_sheet(ws: Worksheet, data: pd.DataFrame, settings: WorkbookSettings) -> None:
    style_title(ws, "Definitions and Data Sources", 8)
    style_section(ws, 3, "Analysis Setup", 1, 8)
    setup = [
        ("Fund name", settings.fund_name),
        ("Input mode", settings.input_mode),
        ("Fund source", settings.fund_source),
        ("Requested start", settings.requested_start),
        ("Requested end", settings.requested_end),
        ("Actual first aligned month", data.index.min().strftime("%Y-%m-%d")),
        ("Actual last aligned month", data.index.max().strftime("%Y-%m-%d")),
        ("Risk-free rate", settings.risk_free_rate_annual),
        ("Crisis definition", f"Month-end VIX level > {settings.crisis_threshold}"),
        ("Rolling regression window", settings.rolling_window),
        ("Uploaded return scale detection", settings.return_scale_method or "N/A"),
    ]
    for row_idx, (label, value) in enumerate(setup, start=4):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
        if label == "Risk-free rate":
            ws.cell(row_idx, 2).number_format = PCT2_FMT

    style_section(ws, 17, "Default Factor Definitions", 1, 8)
    headers = ["Exposure", "Ticker", "Method", "Description", "Source"]
    for col, header in enumerate(headers, start=1):
        ws.cell(18, col, header)
    style_header_row(ws, 18, 1, len(headers))
    for row_idx, factor in enumerate(settings.factors, start=19):
        ws.cell(row_idx, 1, factor.name)
        ws.cell(row_idx, 2, factor.ticker)
        ws.cell(row_idx, 3, factor.method)
        ws.cell(row_idx, 4, factor.description)
        ws.cell(row_idx, 5, "Yahoo Finance via yfinance")
        imported_cell(ws.cell(row_idx, 5))

    style_section(ws, 26, "Statistical Conventions", 1, 8)
    conventions = [
        "OLS means ordinary least squares: coefficients minimize squared residuals.",
        "Standard regression uses the four selected factor returns.",
        "Conditional regression adds crisis dummy x factor interaction terms.",
        "Volatility factor is monthly percentage change in VIX; crisis dummy uses month-end VIX level.",
        "Annual arithmetic return = monthly arithmetic average return x 12.",
        "Annualized standard deviation = monthly standard deviation x SQRT(12).",
        "Sharpe and Sortino subtract the user-selected annual risk-free rate from annualized arithmetic return.",
        "Annual historical VaR/CVaR use rolling 12-month compounded returns, not parametric scaling of monthly VaR.",
        "Drawdown is based on monthly return resolution.",
        "Upside/downside capture use average monthly fund return divided by average monthly benchmark return in months where benchmark return is above/below zero.",
    ]
    for row_idx, text in enumerate(conventions, start=27):
        ws.cell(row_idx, 1, text)

    style_section(ws, 40, "Data Source Notes", 1, 8)
    notes = [
        "Public security total returns are calculated from Yahoo Finance adjusted close, which is intended to reflect dividends/distributions and splits.",
        "Raw market data is retrieved at runtime by yfinance. The workbook stores the monthly return series used in the model.",
        "All model series are aligned by month-end and incomplete months are dropped before regression.",
    ]
    for row_idx, text in enumerate(notes, start=41):
        ws.cell(row_idx, 1, text)

    set_widths(ws, {"A": 34, "B": 28, "C": 18, "D": 74, "E": 28, "F": 16, "G": 16, "H": 16})


def _write_summary_sheet(
    ws: Worksheet,
    data: pd.DataFrame,
    settings: WorkbookSettings,
    first_row: int,
    last_row: int,
    has_benchmark: bool,
) -> None:
    style_title(ws, "Factor Risk Analysis Summary", 8)
    style_section(ws, 3, "Analysis Inputs", 1, 5)
    inputs = [
        ("Fund", settings.fund_name),
        ("Input mode", settings.input_mode),
        ("Requested date range", f"{settings.requested_start} to {settings.requested_end}"),
        ("Aligned date range", f"{data.index.min().strftime('%Y-%m-%d')} to {data.index.max().strftime('%Y-%m-%d')}"),
        ("Complete observations", f"=COUNT({xl_range(SHEET_FUND, 2, first_row, last_row)})"),
        ("Benchmark", settings.benchmark_name or "None"),
        ("Risk-free rate", f"={qsheet(SHEET_RISK)}!B4"),
        ("Crisis definition", f"Month-end VIX > {settings.crisis_threshold}"),
        ("Rolling window", settings.rolling_window),
    ]
    for i, (label, value) in enumerate(inputs, start=4):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
        if isinstance(value, str) and value.startswith("="):
            formula_cell(ws.cell(i, 2))
        else:
            static_cell(ws.cell(i, 2))
    ws["B10"].number_format = PCT2_FMT

    style_section(ws, 15, "Regression Summary", 1, 5)
    headers = ["Metric", "Standard", "Conditional"]
    for col, header in enumerate(headers, start=1):
        ws.cell(16, col, header)
    style_header_row(ws, 16, 1, len(headers))
    rows = [
        ("R-squared", f"={qsheet(SHEET_REGRESSION)}!B24", f"={qsheet(SHEET_CONDITIONAL)}!B24", PCT2_FMT),
        ("Adjusted R-squared", f"={qsheet(SHEET_REGRESSION)}!B25", f"={qsheet(SHEET_CONDITIONAL)}!B25", PCT2_FMT),
        ("F-statistic", f"={qsheet(SHEET_REGRESSION)}!B26", f"={qsheet(SHEET_CONDITIONAL)}!B26", NUM_FMT),
        ("Model p-value", f"={qsheet(SHEET_REGRESSION)}!B27", f"={qsheet(SHEET_CONDITIONAL)}!B27", PVALUE_FMT),
    ]
    for row_idx, (label, std_formula, cond_formula, numfmt) in enumerate(rows, start=17):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, std_formula)
        ws.cell(row_idx, 3, cond_formula)
        for col in (2, 3):
            ws.cell(row_idx, col).number_format = numfmt
            formula_cell(ws.cell(row_idx, col))

    style_section(ws, 23, "Selected Risk Metrics", 1, 5)
    headers = ["Metric", settings.fund_name or "Fund", settings.benchmark_name or "Benchmark"]
    for col, header in enumerate(headers, start=1):
        ws.cell(24, col, header)
    style_header_row(ws, 24, 1, len(headers))
    metric_refs = [
        ("Annualized arithmetic return", "B13", "C13", PCT2_FMT),
        ("Annualized standard deviation", "B15", "C15", PCT2_FMT),
        ("Sharpe ratio", "B18", "C18", NUM_FMT),
        ("Sortino ratio", "B19", "C19", NUM_FMT),
        ("Maximum drawdown", "B20", "C20", PCT2_FMT),
        ("Annual 95% historical VaR", "B21", "C21", PCT2_FMT),
        ("Annual 95% historical CVaR", "B22", "C22", PCT2_FMT),
    ]
    for row_idx, (label, fund_ref, bench_ref, numfmt) in enumerate(metric_refs, start=25):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, f"={qsheet(SHEET_RISK)}!{fund_ref}")
        ws.cell(row_idx, 3, f"={qsheet(SHEET_RISK)}!{bench_ref}" if has_benchmark else "")
        for col in (2, 3):
            ws.cell(row_idx, col).number_format = numfmt
            formula_cell(ws.cell(row_idx, col))

    style_section(ws, 34, "Warnings", 1, 5)
    if settings.data_warnings:
        for row_idx, warning in enumerate(settings.data_warnings, start=35):
            ws.cell(row_idx, 1, warning)
            ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor=WARNING_FILL)
    else:
        ws["A35"] = "No warnings."

    set_widths(ws, {"A": 34, "B": 22, "C": 22, "D": 14, "E": 14, "F": 4, "G": 20, "H": 20, "I": 20, "J": 20})


def _add_summary_charts(wb: Workbook, data: pd.DataFrame, first_row: int, last_row: int, has_benchmark: bool) -> None:
    ws_summary = wb[SHEET_SUMMARY]
    ws_fund = wb[SHEET_FUND]
    ws_reg = wb[SHEET_REGRESSION]
    ws_roll = wb[SHEET_ROLLING]

    # Cumulative return / wealth index.
    chart = LineChart()
    chart.title = "Cumulative Wealth Index"
    chart.y_axis.title = "Index"
    chart.x_axis.title = "Month"
    max_col = 7 if has_benchmark else 4
    # Non-contiguous references are not supported by openpyxl charts, so add series separately.
    chart.add_data(Reference(ws_fund, min_col=4, min_row=4, max_row=last_row), titles_from_data=True)
    if has_benchmark:
        chart.add_data(Reference(ws_fund, min_col=7, min_row=4, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws_fund, min_col=1, min_row=first_row, max_row=last_row))
    chart.height = 8
    chart.width = 15
    ws_summary.add_chart(chart, "G3")

    drawdown = LineChart()
    drawdown.title = "Drawdown"
    drawdown.y_axis.title = "Drawdown"
    drawdown.x_axis.title = "Month"
    drawdown.add_data(Reference(ws_fund, min_col=6, min_row=4, max_row=last_row), titles_from_data=True)
    if has_benchmark:
        drawdown.add_data(Reference(ws_fund, min_col=9, min_row=4, max_row=last_row), titles_from_data=True)
    drawdown.set_categories(Reference(ws_fund, min_col=1, min_row=first_row, max_row=last_row))
    drawdown.height = 8
    drawdown.width = 15
    ws_summary.add_chart(drawdown, "G19")

    avp_start = 31
    avp_end = 31 + len(data)
    avp = LineChart()
    avp.title = "Actual vs. Predicted Monthly Returns"
    avp.y_axis.title = "Return"
    avp.x_axis.title = "Month"
    avp.add_data(Reference(ws_reg, min_col=2, min_row=avp_start, max_row=avp_end), titles_from_data=True)
    avp.add_data(Reference(ws_reg, min_col=3, min_row=avp_start, max_row=avp_end), titles_from_data=True)
    avp.set_categories(Reference(ws_reg, min_col=1, min_row=avp_start + 1, max_row=avp_end))
    avp.height = 8
    avp.width = 15
    ws_summary.add_chart(avp, "G35")

    beta_chart = BarChart()
    beta_chart.title = "Standard Model Factor Exposures"
    beta_chart.y_axis.title = "Beta"
    beta_chart.x_axis.title = "Factor"
    beta_chart.add_data(Reference(ws_reg, min_col=2, min_row=15, max_row=18), titles_from_data=False)
    beta_chart.set_categories(Reference(ws_reg, min_col=1, min_row=15, max_row=18))
    beta_chart.height = 8
    beta_chart.width = 15
    ws_summary.add_chart(beta_chart, "G51")

    rolling_end = first_row + len(data) - 1
    rolling_beta = LineChart()
    rolling_beta.title = "Rolling Factor Exposures"
    rolling_beta.y_axis.title = "Beta"
    rolling_beta.x_axis.title = "Month"
    rolling_beta.add_data(Reference(ws_roll, min_col=2, max_col=5, min_row=4, max_row=rolling_end), titles_from_data=True)
    rolling_beta.set_categories(Reference(ws_roll, min_col=1, min_row=first_row, max_row=rolling_end))
    rolling_beta.height = 8
    rolling_beta.width = 15
    ws_summary.add_chart(rolling_beta, "W3")

    rolling_r2 = LineChart()
    rolling_r2.title = "Rolling R-squared"
    rolling_r2.y_axis.title = "R-squared"
    rolling_r2.x_axis.title = "Month"
    rolling_r2.add_data(Reference(ws_roll, min_col=6, min_row=4, max_row=rolling_end), titles_from_data=True)
    rolling_r2.set_categories(Reference(ws_roll, min_col=1, min_row=first_row, max_row=rolling_end))
    rolling_r2.height = 8
    rolling_r2.width = 15
    ws_summary.add_chart(rolling_r2, "W19")
